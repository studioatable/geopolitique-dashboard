#!/usr/bin/env python3
"""
sipri_milex.py — Ingestion SIPRI Military Expenditure Database
==============================================================

Télécharge le fichier XLSX SIPRI MILEX, parse les feuilles "Constant USD"
et "Share of GDP", normalise en JSON multi-indicateurs.

Approche v0.4.0 :
    - URL du fichier hardcodée (révision annuelle SIPRI prévisible — typiquement
      avril). Évolution prévue v0.4.1 : scraping de la page d'accueil
      https://www.sipri.org/databases/milex pour récupérer dynamiquement le
      lien du fichier le plus récent.
    - Auto-détection de la ligne d'en-tête (recherche d'une ligne contenant
      au moins 5 années consécutives), pour résister aux changements de
      mise en forme entre révisions SIPRI.
    - Auto-détection des feuilles cibles par nom approximatif (le libellé
      exact varie selon l'année de référence USD constants).

Sortie :
    - data/raw/sipri-milex-latest.xlsx (XLSX brut, écrasé)
    - data/defense/sipri_milex.json (JSON normalisé multi-indicateurs)

Usage :
    python ingestion/sipri_milex.py

Code de sortie :
    0 — ingestion réussie
    1 — erreur réseau, parsing ou feuille absente

Studio à Table — geopolitique-dashboard
"""

from __future__ import annotations

import io
import json
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import openpyxl
import requests

SCRIPT_VERSION = "0.4.0"
SOURCE_ID = "sipri_milex"
TIMEOUT_SECONDS = 60
USER_AGENT = (
    f"Studio-a-Table-Geopolitique-Ingest/{SCRIPT_VERSION} "
    "(+https://studioatable.fr)"
)

# URL du fichier XLSX SIPRI MILEX, révision avril 2026.
# À mettre à jour annuellement (ou via scraping en v0.4.1).
XLSX_URL = (
    "https://www.sipri.org/sites/default/files/"
    "SIPRI-Milex-data-1949-2025_v1.2.xlsx"
)

# Indicateurs cibles avec patterns de détection des feuilles SIPRI.
# Les libellés exacts varient selon la révision (année de référence USD).
#
# `name_alternatives` est une liste d'alternatives ; chaque alternative est
# elle-même une liste de mots-clés qui doivent TOUS être présents dans le nom
# de feuille (lower-case). La feuille est matchée dès la première alternative
# qui passe — sémantique : OR entre alternatives, AND à l'intérieur de chacune.
SHEET_TARGETS = [
    {
        "indicator_id": "milex_constant_usd",
        "unit": "USD constants (millions)",
        "value_multiplier": 1.0,
        "name_alternatives": [
            ["constant", "us$"],  # ex. "Constant (2024) US$"
            ["constant", "usd"],
        ],
    },
    {
        "indicator_id": "milex_pct_gdp",
        "unit": "% du PIB",
        # SIPRI stocke "Share of GDP" en ratio décimal (0,0194 = 1,94 %).
        # On multiplie par 100 pour que la valeur dans le JSON soit déjà
        # exprimée en pourcentage humain (1,94), conforme à l'unité.
        "value_multiplier": 100.0,
        "name_alternatives": [
            ["share", "gdp"],     # ex. "Share of GDP"
            ["% of gdp"],
        ],
    },
]

# Marqueurs "donnée indisponible" rencontrés dans les fichiers SIPRI
MISSING_MARKERS = {None, "", "xxx", "...", ". .", "..", " ", "  "}

ROOT = Path(__file__).resolve().parent.parent
REGISTRY_PATH = ROOT / "data" / "registry" / "sources.json"
RAW_DIR = ROOT / "data" / "raw"
DEFENSE_DIR = ROOT / "data" / "defense"
RAW_OUTPUT = RAW_DIR / "sipri-milex-latest.xlsx"
JSON_OUTPUT = DEFENSE_DIR / "sipri_milex.json"


def load_source_meta() -> dict:
    """Charge la fiche SIPRI MILEX du registre local."""
    if not REGISTRY_PATH.exists():
        raise FileNotFoundError(
            f"Registre absent : {REGISTRY_PATH}\n"
            "Copier sources_geopolitiques.json depuis le dossier parent."
        )
    with REGISTRY_PATH.open("r", encoding="utf-8") as f:
        registry = json.load(f)
    for s in registry.get("sources", []):
        if s.get("id") == SOURCE_ID:
            return s
    raise KeyError(f"Source {SOURCE_ID} absente du registre.")


def fetch_xlsx(url: str) -> bytes:
    """Télécharge le fichier XLSX SIPRI."""
    headers = {"User-Agent": USER_AGENT, "Accept": "*/*"}
    resp = requests.get(url, headers=headers, timeout=TIMEOUT_SECONDS)
    resp.raise_for_status()
    return resp.content


def match_sheet(
    sheet_names: list[str], alternatives: list[list[str]]
) -> Optional[str]:
    """Cherche la première feuille correspondant à au moins une alternative.

    Chaque alternative est une liste de mots-clés qui doivent TOUS être
    présents (en lower-case) dans le nom de la feuille pour qu'elle matche.
    Une feuille passe dès la première alternative satisfaite.
    """
    for name in sheet_names:
        lc_name = name.lower()
        for alt_keywords in alternatives:
            if all(kw.lower() in lc_name for kw in alt_keywords):
                return name
    return None


def find_header_row(ws) -> tuple[int, list[int], list[int]]:
    """Détecte la ligne d'en-tête en cherchant la 1ère ligne contenant
    au moins 5 années consécutives en colonnes (1940-2100).

    Retourne : (index_ligne, indices_colonnes_années, valeurs_années).
    """
    max_scan_rows = min(20, ws.max_row)
    for row_idx in range(1, max_scan_rows + 1):
        years: list[int] = []
        year_cols: list[int] = []
        for col_idx in range(1, ws.max_column + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if isinstance(value, int) and 1940 < value < 2100:
                years.append(value)
                year_cols.append(col_idx)
            elif isinstance(value, float) and value.is_integer() and 1940 < int(value) < 2100:
                years.append(int(value))
                year_cols.append(col_idx)
        if len(years) >= 5:
            return row_idx, year_cols, years
    raise ValueError(
        "Impossible de localiser la ligne d'en-tête (>= 5 années consécutives)."
    )


def is_data_value(value) -> bool:
    """True si la cellule contient une donnée numérique exploitable."""
    if value in MISSING_MARKERS:
        return False
    return isinstance(value, (int, float))


def is_country_label(value) -> bool:
    """True si la cellule semble être un nom de pays exploitable."""
    if not isinstance(value, str):
        return False
    stripped = value.strip()
    if not stripped:
        return False
    # Exclure lignes de notes, totaux, séparateurs
    lower = stripped.lower()
    blacklist_prefixes = (
        "notes", "source", "table", "* ", "** ", "world total",
        "africa", "americas", "asia", "europe", "middle east", "oceania",
        # Régions agrégées : on les exclut du périmètre pays. À reconsidérer si
        # on veut intégrer les agrégats régionaux SIPRI dans un futur indicateur dédié.
    )
    if any(lower.startswith(p) for p in blacklist_prefixes):
        return False
    return True


def extract_sheet(ws, indicator: dict) -> dict:
    """Extrait les données d'une feuille selon la grammaire SIPRI :
    colonne A = pays, colonnes suivantes = années."""
    header_row, year_cols, years = find_header_row(ws)
    multiplier = float(indicator.get("value_multiplier", 1.0))

    countries_data: dict[str, dict[str, float]] = {}
    for row_idx in range(header_row + 1, ws.max_row + 1):
        country_value = ws.cell(row=row_idx, column=1).value
        if not is_country_label(country_value):
            continue
        country_name = country_value.strip()

        year_data: dict[str, float] = {}
        for year, col_idx in zip(years, year_cols):
            value = ws.cell(row=row_idx, column=col_idx).value
            if not is_data_value(value):
                continue
            year_data[str(year)] = round(float(value) * multiplier, 6)

        if year_data:
            countries_data[country_name] = year_data

    return {
        "indicator_id": indicator["indicator_id"],
        "unit": indicator["unit"],
        "value_multiplier_applied": multiplier,
        "years_min": min(years) if years else None,
        "years_max": max(years) if years else None,
        "countries_count": len(countries_data),
        "data": countries_data,
    }


def write_outputs(
    raw_xlsx: bytes,
    indicators: list[dict],
    source_meta: dict,
    feed_url: str,
    notes: list[str],
) -> None:
    """Écrit le brut + le JSON normalisé."""
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    DEFENSE_DIR.mkdir(parents=True, exist_ok=True)

    RAW_OUTPUT.write_bytes(raw_xlsx)

    now_utc = datetime.now(timezone.utc).isoformat()
    payload = {
        "schema_version": "1.0.0",
        "ingestion_script": "ingestion/sipri_milex.py",
        "script_version": SCRIPT_VERSION,
        "source": {
            "id": SOURCE_ID,
            "name": source_meta.get("short_name", "SIPRI MILEX"),
            "provider": source_meta.get(
                "provider", "Stockholm International Peace Research Institute"
            ),
            "category": source_meta.get("category", "defense"),
            "xlsx_url": feed_url,
            "reliability": source_meta.get("reliability"),
            "license": source_meta.get("license"),
        },
        "fetched_at_utc": now_utc,
        "indicators": {ind["indicator_id"]: ind for ind in indicators},
        "indicators_count": len(indicators),
        "notes": notes,
    }
    with JSON_OUTPUT.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def main() -> int:
    print(f"=== Ingestion SIPRI MILEX — script {SCRIPT_VERSION} ===")

    try:
        source_meta = load_source_meta()
    except (FileNotFoundError, KeyError) as e:
        print(f"FATAL — {e}", file=sys.stderr)
        return 2

    print(f"Fichier ciblé : {XLSX_URL}")

    try:
        raw_xlsx = fetch_xlsx(XLSX_URL)
    except requests.RequestException as e:
        print(f"ERREUR réseau : {e.__class__.__name__} — {e}", file=sys.stderr)
        return 1

    print(f"XLSX téléchargé : {len(raw_xlsx):,} octets")

    try:
        # data_only=True : on récupère les valeurs calculées plutôt que les
        # formules. read_only retiré : il est mal compatible avec l'accès
        # indexé ws.cell(row=, column=) et provoque des lenteurs majeures sur
        # ce type de parsing. Le fichier SIPRI fait ~1 Mo, charger en mémoire
        # est sans coût significatif.
        wb = openpyxl.load_workbook(io.BytesIO(raw_xlsx), data_only=True)
    except Exception as e:
        print(f"ERREUR parsing XLSX : {e}", file=sys.stderr)
        return 1

    sheet_names = wb.sheetnames
    print(f"Feuilles disponibles : {sheet_names}")

    indicators: list[dict] = []
    notes: list[str] = []

    for target in SHEET_TARGETS:
        matched = match_sheet(sheet_names, target["name_alternatives"])
        if matched is None:
            note = (
                f"Feuille {target['indicator_id']} introuvable "
                f"(alternatives {target['name_alternatives']}). Ignoré."
            )
            print(f"  ! {note}")
            notes.append(note)
            continue

        print(f"  → Extraction feuille « {matched} » ...", end=" ", flush=True)
        try:
            ws = wb[matched]
            indicator_data = extract_sheet(ws, target)
            indicator_data["sheet_source"] = matched
            indicators.append(indicator_data)
            print(
                f"OK ({indicator_data['countries_count']} pays, "
                f"{indicator_data['years_min']}-{indicator_data['years_max']})"
            )
        except Exception as e:
            note = f"Erreur extraction « {matched} » : {e}"
            print(f"ERR — {note}")
            notes.append(note)

    if not indicators:
        print("FATAL — aucun indicateur extrait.", file=sys.stderr)
        return 1

    write_outputs(raw_xlsx, indicators, source_meta, XLSX_URL, notes)

    print(f"\nBrut écrit : {RAW_OUTPUT.relative_to(ROOT)}")
    print(f"JSON écrit : {JSON_OUTPUT.relative_to(ROOT)}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
